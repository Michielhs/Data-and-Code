import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook

# Load new data
file_path_ex_ante_new = "Data results/GPT4/RUN5/RUN5.csv_comparison_ex_ante.csv"
file_path_ad_interim_new = "Data results/GPT4/RUN5/RUN5.csv_comparison_ad_interim.csv"

# Load original data
file_path_ex_ante = "Input data/experimental_data_flip_a_coin.csv_comparison_ex_ante.csv"
file_path_ad_interim = "Input data/experimental_data_flip_a_coin.csv_comparison_ad_interim.csv"

# set name of results
file_path = 'ResultsP1.xlsx'

# Load new data from CSV
def load_data(filename):
    return pd.read_csv(filename)

# Prepare data by calculating the percentage for each comparison and converting to real percentages
def prep_data(df):
    df = df.copy()  # Ensure df is a copy if not already
    conditions = [
        df['BinaryChoice'] == "AGV vs. SM",
        df['BinaryChoice'] == "AGV vs. NSQ",
        df['BinaryChoice'] == "AGV vs. RAND",
        df['BinaryChoice'] == "SM vs. NSQ",
        df['BinaryChoice'] == "SM vs. RAND",
        df['BinaryChoice'] == "NSQ vs. RAND"
    ]
    choices = ['chose_AGV', 'chose_AGV', 'chose_AGV', 'chose_SM', 'chose_SM', 'chose_NSQ']
    percentages = np.select(conditions, [df[c] / df['total'] * 100 for c in choices], default=-1)
    df.loc[:, 'percentage'] = percentages
    return df

# Generate table of percentages with multi-level header
def generate_table(df):
    pivot_table = df.pivot_table(index=['treatment_distribution'], columns='BinaryChoice', values='percentage', aggfunc='mean')
    # Add binary choice in the first row and choices in the second row
    columns = pd.MultiIndex.from_tuples([(col, 'percentage') for col in pivot_table.columns])
    pivot_table.columns = columns
    pivot_table = pivot_table.reset_index()
    return pivot_table

# Create a new workbook and clear existing data
def create_new_workbook(file_path, sheet_names):
    wb = Workbook()
    for sheet_name in sheet_names:
        wb.create_sheet(title=sheet_name)
    del wb['Sheet']  # Remove the default sheet created by openpyxl
    wb.save(file_path)

# Function to append DataFrame to an existing Excel sheet using openpyxl
def append_df_to_excel(file_path, df, sheet_name, startrow=None, empty_rows=0):
    # Load existing workbook
    book = load_workbook(file_path)

    if sheet_name not in book.sheetnames:
        raise ValueError(f"Sheet {sheet_name} does not exist in {file_path}")

    sheet = book[sheet_name]

    if startrow is None:
        startrow = sheet.max_row

    # Insert empty rows if needed
    for _ in range(empty_rows):
        sheet.append([])

    # Append the DataFrame with multi-level header
    for row in pd.concat([pd.DataFrame(df.columns.tolist()).T, df]).itertuples(index=False, name=None):
        sheet.append(row)

    book.save(file_path)

# Function to write the statistics to a new sheet
def write_statistics_to_excel(file_path, statistics_df, sheet_name):
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        statistics_df.to_excel(writer, sheet_name=sheet_name, index=False)

# Create a new workbook and clear existing data
sheet_names = ['Ex Ante', 'Ad Interim Val Pos', 'Ad Interim Val Neg']
create_new_workbook(file_path, sheet_names)


# Load data
data_ex_ante = load_data(file_path_ex_ante)
data_ad_interim = load_data(file_path_ad_interim)

# Filter data for ex_ante based on criteria (e.g., rounds)
data_ex_ante_filtered = data_ex_ante[data_ex_ante['ad_interim_round'] == 0].copy()

# Filter data for ad_interim based on criteria (e.g., rounds and valuation_positive)
data_ad_interim_val_pos = data_ad_interim[(data_ad_interim['ad_interim_round'] == 1) & (data_ad_interim['valuation_positive'] == 1)].copy()
data_ad_interim_val_neg = data_ad_interim[(data_ad_interim['ad_interim_round'] == 1) & (data_ad_interim['valuation_positive'] == 0)].copy()

# Prepare and generate tables
data_ex_ante_prepared = prep_data(data_ex_ante_filtered)
table_ex_ante = generate_table(data_ex_ante_prepared)

data_ad_interim_val_pos_prepared = prep_data(data_ad_interim_val_pos)
data_ad_interim_val_neg_prepared = prep_data(data_ad_interim_val_neg)

table_ad_interim_val_pos = generate_table(data_ad_interim_val_pos_prepared)
table_ad_interim_val_neg = generate_table(data_ad_interim_val_neg_prepared)


data_ex_ante_new = load_data(file_path_ex_ante_new)
data_ad_interim_new = load_data(file_path_ad_interim_new)

# Filter new data
data_ex_ante_new_filtered = data_ex_ante_new[data_ex_ante_new['ad_interim_round'] == 0]
data_ad_interim_new_val_pos = data_ad_interim_new[(data_ad_interim_new['ad_interim_round'] == 1) & (data_ad_interim_new['valuation_positive'] == 1)]
data_ad_interim_new_val_neg = data_ad_interim_new[(data_ad_interim_new['ad_interim_round'] == 1) & (data_ad_interim_new['valuation_positive'] == 0)]

# Prepare and generate new tables
data_ex_ante_new_prepared = prep_data(data_ex_ante_new_filtered)
data_ad_interim_new_val_pos_prepared = prep_data(data_ad_interim_new_val_pos)
data_ad_interim_new_val_neg_prepared = prep_data(data_ad_interim_new_val_neg)

table_ex_ante_new = generate_table(data_ex_ante_new_prepared)
table_ad_interim_val_pos_new = generate_table(data_ad_interim_new_val_pos_prepared)
table_ad_interim_val_neg_new = generate_table(data_ad_interim_new_val_neg_prepared)

# Calculate absolute differences
table_ex_ante_diff_abs = (table_ex_ante_new.set_index('treatment_distribution') - table_ex_ante.set_index('treatment_distribution')).abs().reset_index()
table_ad_interim_val_pos_diff_abs = (table_ad_interim_val_pos_new.set_index('treatment_distribution') - table_ad_interim_val_pos.set_index('treatment_distribution')).abs().reset_index()
table_ad_interim_val_neg_diff_abs = (table_ad_interim_val_neg_new.set_index('treatment_distribution') - table_ad_interim_val_neg.set_index('treatment_distribution')).abs().reset_index()

# Calculate squared differences
table_ex_ante_diff_squared = (table_ex_ante_new.set_index('treatment_distribution') - table_ex_ante.set_index('treatment_distribution')) ** 2
table_ex_ante_diff_squared = table_ex_ante_diff_squared.reset_index()
table_ad_interim_val_pos_diff_squared = (table_ad_interim_val_pos_new.set_index('treatment_distribution') - table_ad_interim_val_pos.set_index('treatment_distribution')) ** 2
table_ad_interim_val_pos_diff_squared = table_ad_interim_val_pos_diff_squared.reset_index()
table_ad_interim_val_neg_diff_squared = (table_ad_interim_val_neg_new.set_index('treatment_distribution') - table_ad_interim_val_neg.set_index('treatment_distribution')) ** 2
table_ad_interim_val_neg_diff_squared = table_ad_interim_val_neg_diff_squared.reset_index()

import pandas as pd
import numpy as np
from openpyxl import load_workbook, Workbook

# Compute max, min, mean, and sum values for the differences
def compute_statistics(df, label):
    numeric_df = df.select_dtypes(include=[np.number])  # Select only numeric columns
    return pd.DataFrame({
        'Metric': ['Max', 'Min', 'Mean', 'Sum'],
        'Value': [numeric_df.max().max(), numeric_df.min().min(), numeric_df.mean().mean(), numeric_df.sum().sum()],
        'Category': label
    })

# Calculate statistics for each table of absolute differences
stats_ex_ante_abs = compute_statistics(table_ex_ante_diff_abs, 'Ex Ante (Abs)')
stats_ad_interim_val_pos_abs = compute_statistics(table_ad_interim_val_pos_diff_abs, 'Ad Interim Val Pos (Abs)')
stats_ad_interim_val_neg_abs = compute_statistics(table_ad_interim_val_neg_diff_abs, 'Ad Interim Val Neg (Abs)')

# Calculate statistics for each table of squared differences
stats_ex_ante_squared = compute_statistics(table_ex_ante_diff_squared, 'Ex Ante (Squared)')
stats_ad_interim_val_pos_squared = compute_statistics(table_ad_interim_val_pos_diff_squared, 'Ad Interim Val Pos (Squared)')
stats_ad_interim_val_neg_squared = compute_statistics(table_ad_interim_val_neg_diff_squared, 'Ad Interim Val Neg (Squared)')

# Combine statistics into a single DataFrame
statistics_df_abs = pd.concat([stats_ex_ante_abs, stats_ad_interim_val_pos_abs, stats_ad_interim_val_neg_abs], axis=0)
statistics_df_squared = pd.concat([stats_ex_ante_squared, stats_ad_interim_val_pos_squared, stats_ad_interim_val_neg_squared], axis=0)

# Calculate overall statistics across all absolute differences
overall_stats_abs = statistics_df_abs.groupby('Metric')['Value'].agg(['max', 'min', 'mean', 'sum']).reset_index()
overall_stats_abs.columns = ['Metric', 'Total Max (Abs)', 'Total Min (Abs)', 'Total Mean (Abs)', 'Total Sum (Abs)']

# Calculate overall statistics across all squared differences
overall_stats_squared = statistics_df_squared.groupby('Metric')['Value'].agg(['max', 'min', 'mean', 'sum']).reset_index()
overall_stats_squared.columns = ['Metric', 'Total Max (Squared)', 'Total Min (Squared)', 'Total Mean (Squared)', 'Total Sum (Squared)']

# Combine overall statistics into one DataFrame
overall_stats = pd.merge(overall_stats_abs, overall_stats_squared, on='Metric')

# Combine all statistics
combined_statistics_df = pd.concat([statistics_df_abs, statistics_df_squared], axis=0)
combined_statistics_df = combined_statistics_df.pivot(index='Metric', columns='Category', values='Value').reset_index()

# Add overall statistics to the combined DataFrame
combined_statistics_df = pd.concat([combined_statistics_df, overall_stats.set_index('Metric').T.reset_index().rename(columns={'index': 'Metric'})], ignore_index=True)

# Function to write the statistics to a new sheet
def write_statistics_to_excel(file_path, statistics_df, sheet_name):
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a') as writer:
        statistics_df.to_excel(writer, sheet_name=sheet_name, index=False)

# Write the statistics to a new sheet called "Part 1"
write_statistics_to_excel(file_path, combined_statistics_df, 'Part 1')

# Append the new data to the existing sheets
append_df_to_excel(file_path, table_ex_ante, sheet_name='Ex Ante', empty_rows=0)
append_df_to_excel(file_path, table_ex_ante_new, sheet_name='Ex Ante', empty_rows=0)
append_df_to_excel(file_path, table_ex_ante_diff_abs, sheet_name='Ex Ante', empty_rows=2)
append_df_to_excel(file_path, table_ex_ante_diff_squared, sheet_name='Ex Ante', empty_rows=0)

append_df_to_excel(file_path, table_ad_interim_val_pos, sheet_name='Ad Interim Val Pos', empty_rows=0)
append_df_to_excel(file_path, table_ad_interim_val_pos_new, sheet_name='Ad Interim Val Pos', empty_rows=0)
append_df_to_excel(file_path, table_ad_interim_val_pos_diff_abs, sheet_name='Ad Interim Val Pos', empty_rows=2)
append_df_to_excel(file_path, table_ad_interim_val_pos_diff_squared, sheet_name='Ad Interim Val Pos', empty_rows=0)

append_df_to_excel(file_path, table_ad_interim_val_neg, sheet_name='Ad Interim Val Neg', empty_rows=0)
append_df_to_excel(file_path, table_ad_interim_val_neg_new, sheet_name='Ad Interim Val Neg', empty_rows=0)
append_df_to_excel(file_path, table_ad_interim_val_neg_diff_abs, sheet_name='Ad Interim Val Neg', empty_rows=2)
append_df_to_excel(file_path, table_ad_interim_val_neg_diff_squared, sheet_name='Ad Interim Val Neg', empty_rows=0)

print(f"Tables and statistics have been appended to {file_path}")